#!/usr/bin/env python3
"""Monitor de estado de anuncios del shortlist (Inmuebles_SHORTLIST.xlsx).

Para cada fila con URL en la hoja `Inmuebles`:
  1. Abre la URL en Playwright (perfil persistente por portal, resuelve captcha en headed).
  2. Detecta el estado del anuncio: Activo / Retirado / Caducado / Vendido / Reservado / Desconocido.
  3. Compara el precio actual contra `Precio publicado` (col 12) y anota la diferencia.
  4. Rellena oportunisticamente Municipio (col 9), Barrio (col 10) y m2 anuncio (col 18) si estan vacios.
  5. Actualiza `Ultima actualizacion` (col 4) y escribe el estado en una nueva columna `Estado anuncio`.

Reutiliza `scrape_listing.py` como modulo. Edita el .xlsx in-place con openpyxl
para preservar formulas y formato condicional.
"""
from __future__ import annotations

import argparse
import datetime as _dt
import re
import sys
import time
from pathlib import Path

# Reusar el scraper como modulo (mismo directorio).
sys.path.insert(0, str(Path(__file__).resolve().parent))
import scrape_listing as sl  # noqa: E402

DEFAULT_XLSX = (
    Path.home()
    / "Library/CloudStorage/GoogleDrive-rjaumecatany@gmail.com"
    / "My Drive/my_folder/01_Family/Mortgage/Inmuebles/Inmuebles_SHORTLIST.xlsx"
)

# Mapeo de columnas (1-indexed, segun la hoja actual).
COL = {
    "ID": 1,
    "FECHA_PUB": 2,
    "FECHA_DESC": 3,
    "ULTIMA_ACT": 4,
    "PORTAL": 5,
    "REF": 6,
    "URL": 7,
    "MUNICIPIO": 9,
    "BARRIO": 10,
    "PRECIO": 12,
    "M2": 18,
    "ESTADO_PROCESO": 38,
}
# Columna nueva: se decide en runtime (max_column + 1 si no existe ya).
NEW_COL_HEADER = "Estado anuncio"

# Indicadores de off-market (HTML en minuscula).
OFF_MARKET_PATTERNS = [
    (re.compile(r"anuncio\s+retirado"), "Retirado"),
    (re.compile(r"este\s+anuncio\s+ya\s+no\s+(esta|está)\s+disponible"), "Retirado"),
    (re.compile(r"este\s+anuncio\s+ha\s+sido\s+eliminado"), "Retirado"),
    (re.compile(r"anuncio\s+(caducado|expirado)"), "Caducado"),
    (re.compile(r"(este\s+anuncio|esta\s+vivienda|este\s+inmueble).{0,80}\bvendid[oa]\b", re.DOTALL), "Vendido"),
    (re.compile(r"(este\s+anuncio|esta\s+vivienda|este\s+inmueble).{0,80}\breservad[oa]\b", re.DOTALL), "Reservado"),
]


def detect_listing_status(html: str, portal: str, data: dict) -> str:
    """Devuelve 'Activo'|'Retirado'|'Caducado'|'Vendido'|'Reservado'|'Desconocido'."""
    if not html:
        return "Desconocido"
    txt = html.lower()
    for rx, label in OFF_MARKET_PATTERNS:
        if rx.search(txt):
            return label
    # Si hay precio extraido, asumimos activo.
    if data.get("precio_publicado"):
        return "Activo"
    return "Desconocido"


def _parse_eur(s) -> int | None:
    """Toma el PRIMER numero estilo europeo de un string y lo devuelve como int.

    '279.000 EUR' -> 279000
    '279.000 €' -> 279000
    '2.300.000 € 2.225.000 €' -> 2300000   (toma solo el primero; usar _parse_eur_all si necesitas mas)
    '215000' -> 215000
    """
    if s is None:
        return None
    if isinstance(s, (int, float)):
        return int(s)
    text = str(s).strip()
    # 1) Numero europeo con separadores: "279.000"
    m = re.search(r"\d{1,3}(?:[. \s]\d{3})+", text)
    # 2) Fallback: 5+ digitos pegados (evita matches de refs cortas como "9035")
    if not m:
        m = re.search(r"\d{5,}", text)
    if not m:
        return None
    digits = re.sub(r"[^\d]", "", m.group(0))
    return int(digits) if digits else None


def _fmt_eur(n: int) -> str:
    return f"{n:,}".replace(",", ".") + " EUR"


def _extract_m2(data: dict) -> int | None:
    feats = data.get("caracteristicas") or []
    for f in feats:
        m = re.search(r"(\d{2,4})\s*m[²2]", f)
        if m:
            return int(m.group(1))
    return None


def _profile_dir(portal: str) -> Path:
    """Perfil persistente de Chromium por portal (cookies anti-DataDome)."""
    return Path.home() / ".cache" / f"{portal.lower()}_profile"


def _is_discarded(estado_proceso: str | None) -> bool:
    if not estado_proceso:
        return False
    return "descartado" in str(estado_proceso).lower()


def open_browser_session(portal: str, headed: bool):
    """Devuelve (fetch_url, close) usando un BrowserContext persistente."""
    from playwright.sync_api import sync_playwright

    user_data_dir = str(_profile_dir(portal))
    Path(user_data_dir).mkdir(parents=True, exist_ok=True)

    p = sync_playwright().start()
    ctx = p.chromium.launch_persistent_context(
        user_data_dir,
        headless=not headed,
        locale="es-ES",
        user_agent=sl.DEFAULT_UA,
        args=["--disable-blink-features=AutomationControlled"],
    )

    def fetch(url: str, wait: float = 3.0, timeout: int = 45) -> str:
        page = ctx.new_page()
        try:
            page.goto(url, timeout=timeout * 1000, wait_until="domcontentloaded")
            time.sleep(wait)
            if sl._page_has_captcha(page):
                if headed:
                    sl._wait_for_captcha_resolution(page, max_wait=120)
                else:
                    print(f"  AVISO: captcha en {url} (sin --headed no se puede resolver)", file=sys.stderr)
            return page.content()
        finally:
            page.close()

    def close():
        try:
            ctx.close()
        finally:
            p.stop()

    return fetch, close


def process_row(ws, row_idx: int, fetchers: dict, headed: bool, dry_run: bool, verbose: bool) -> dict:
    """Procesa una fila. Devuelve dict con cambios propuestos/aplicados para reporte."""
    listing_id = ws.cell(row=row_idx, column=COL["ID"]).value
    url = ws.cell(row=row_idx, column=COL["URL"]).value
    if not url:
        return {"id": listing_id, "skip": "sin URL"}

    portal_cell = ws.cell(row=row_idx, column=COL["PORTAL"]).value or ""
    portal_detected, ref = sl.detect_portal_and_ref(url)
    # Si la celda contradice, confiamos en la URL.
    portal = portal_detected if portal_detected != "Desconocido" else (portal_cell or "Desconocido")

    # Obtener fetcher para ese portal (reusa contexto).
    fetcher = fetchers.get(portal)
    if fetcher is None:
        fetch, close = open_browser_session(portal, headed)
        fetchers[portal] = (fetch, close)
        fetcher = fetchers[portal]
    fetch, _ = fetcher

    if verbose:
        print(f"[{listing_id}] GET {url}")
    try:
        html = fetch(url)
    except Exception as e:
        return {"id": listing_id, "error": str(e)}

    data = sl.extract_listing_data(html, url, portal, ref)
    status = detect_listing_status(html, portal, data)

    changes = {"id": listing_id, "url": url, "portal": portal, "status": status, "writes": {}}

    # Precio: comparar con la celda.
    new_price = _parse_eur(data.get("precio_publicado"))
    sheet_price = _parse_eur(ws.cell(row=row_idx, column=COL["PRECIO"]).value)
    if new_price is not None and sheet_price is not None and new_price != sheet_price:
        delta = new_price - sheet_price
        sign = "▼" if delta < 0 else "▲"
        status = f"{status} (precio {sign} {_fmt_eur(sheet_price)} → {_fmt_eur(new_price)})"
        changes["status"] = status
        changes["writes"][COL["PRECIO"]] = _fmt_eur(new_price)

    # Estado anuncio (siempre se escribe).
    changes["writes"][changes.setdefault("_estado_col", None) or 0] = None  # placeholder; rellenamos abajo

    # Ultima actualizacion -> hoy.
    today_str = _dt.date.today().strftime("%Y-%m-%d")
    changes["writes"][COL["ULTIMA_ACT"]] = today_str

    # Re-fill oportunista (solo si vacio).
    municipio_now = ws.cell(row=row_idx, column=COL["MUNICIPIO"]).value
    barrio_now = ws.cell(row=row_idx, column=COL["BARRIO"]).value
    m2_now = ws.cell(row=row_idx, column=COL["M2"]).value
    if (not municipio_now) or (not barrio_now):
        municipio, barrio = sl.derive_location(portal, url, data)
        if not municipio_now and municipio != "Unknown":
            changes["writes"][COL["MUNICIPIO"]] = municipio
        if not barrio_now and barrio != "Unknown":
            changes["writes"][COL["BARRIO"]] = barrio
    if not m2_now:
        m2_val = _extract_m2(data)
        if m2_val:
            changes["writes"][COL["M2"]] = m2_val

    # Limpiar placeholder y guardar status final por separado.
    changes["writes"].pop(0, None)
    changes["status_final"] = status
    return changes


def main() -> int:
    ap = argparse.ArgumentParser(description="Monitor de estado de los anuncios del shortlist.")
    ap.add_argument("--xlsx", default=str(DEFAULT_XLSX), help="Ruta al Inmuebles_SHORTLIST.xlsx")
    ap.add_argument("--only-ids", help="Lista CSV de IDs a procesar (ej. P001,P005). Por defecto todas.")
    ap.add_argument("--include-discarded", action="store_true",
                    help="Procesar tambien filas con 'Estado proceso' que empiezan por DESCARTADO.")
    ap.add_argument("--dry-run", action="store_true", help="No guarda cambios; solo imprime.")
    ap.add_argument("--headed", action="store_true", default=True,
                    help="Playwright visible (default: True; necesario para captchas).")
    ap.add_argument("--headless", dest="headed", action="store_false",
                    help="Forzar headless (NO recomendado en Idealista).")
    ap.add_argument("-v", "--verbose", action="store_true")
    args = ap.parse_args()

    from openpyxl import load_workbook  # local import

    xlsx_path = Path(args.xlsx)
    if not xlsx_path.exists():
        print(f"ERROR: no existe {xlsx_path}", file=sys.stderr)
        return 2

    wb = load_workbook(xlsx_path)
    ws = wb["Inmuebles"]

    # Localizar o crear la columna 'Estado anuncio'.
    headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    if NEW_COL_HEADER in headers:
        estado_col = headers.index(NEW_COL_HEADER) + 1
    else:
        estado_col = ws.max_column + 1
        ws.cell(row=1, column=estado_col).value = NEW_COL_HEADER
    if args.verbose:
        print(f"Columna 'Estado anuncio' = col {estado_col}")

    # Filtrar filas.
    only = set(s.strip() for s in args.only_ids.split(",")) if args.only_ids else None
    rows_to_process = []
    for r in range(2, ws.max_row + 1):
        listing_id = ws.cell(row=r, column=COL["ID"]).value
        if not listing_id:
            continue
        if only and listing_id not in only:
            continue
        if not args.include_discarded and _is_discarded(ws.cell(row=r, column=COL["ESTADO_PROCESO"]).value):
            print(f"[skip {listing_id}: DESCARTADO]")
            continue
        rows_to_process.append((r, listing_id))

    print(f"Filas a procesar: {len(rows_to_process)}")
    print("-" * 80)

    fetchers: dict[str, tuple] = {}
    results = []
    try:
        for r, listing_id in rows_to_process:
            result = process_row(ws, r, fetchers, args.headed, args.dry_run, args.verbose)
            results.append(result)
            if "skip" in result:
                print(f"[skip {listing_id}: {result['skip']}]")
                continue
            if "error" in result:
                print(f"[ERROR {listing_id}: {result['error']}]")
                continue
            status = result["status_final"]
            writes = result["writes"]
            # Resumen por fila
            print(f"[{listing_id}] {status}")
            for col, val in writes.items():
                hdr = headers[col - 1] if col <= len(headers) else "?"
                print(f"      col {col} ({hdr}): -> {val!r}")
            if not args.dry_run:
                ws.cell(row=r, column=estado_col).value = status
                for col, val in writes.items():
                    ws.cell(row=r, column=col).value = val
    finally:
        for portal, (_, close) in fetchers.items():
            try:
                close()
            except Exception:
                pass

    if args.dry_run:
        print("\n--dry-run: no se guarda el .xlsx.")
    else:
        wb.save(xlsx_path)
        print(f"\nGuardado: {xlsx_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
